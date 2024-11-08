import openpyxl
import collections
from datetime import datetime
bank = openpyxl.load_workbook(filename="bank_statement.xlsx")
zoho = openpyxl.load_workbook(filename="zoho_transactions.xlsx")
sheet = bank['s']
z_sheet = zoho['sheet1']
# print(f"{sheet['A2'].value}")





amounts = collections.defaultdict(list)
amounts_zoho = collections.defaultdict(list)



empties = 0
bank_transactions = 0
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    if(not row[1].value):
        empties += 1
        if(empties >= 5):
            break
    else:
        # print(f"{(row[1].value).date() or 'Empty'}:{row[4].value or 'Empty'}")
        amounts[str(row[1].value.date())].append(row[4].value)
        empties = 0
        bank_transactions += 1
print(f"{bank_transactions} Bnk Transactions !")
zoho_transactions = 0
empties = 0
for row in z_sheet.iter_rows(min_row=2, max_row=z_sheet.max_row):
    if(not row[0].value):
        empties += 1
        if(empties >= 5):
            break
    else:
        # print(f"{(row[0].value) or 'Empty'}:{row[4].value or 'Empty'}")
        amounts_zoho[str(row[0].value.date())].append(row[4].value or 'empty')
        empties = 0
        zoho_transactions += 1
    
print(f"{zoho_transactions} zoho transactions!")

choice = 1

choice = int(input("1. Query date\n0.exit\n>>>")) 
while choice:
    query_date = input("Which date: ")
    if(query_date == "0"):
        break
    if(len(query_date) == 1):
        query_date = f"2024-09-0{query_date}"
    elif(len(query_date) == 2):
        query_date = f"2024-09-{query_date}"
    else:
        print()
    total = 0
    for amt in amounts[query_date]:
        print(f"{amt}")
        total += amt
    print(f"Bank Total({query_date}): {total}")

    # print(amounts_zoho)
    total = 0
    for amt in amounts_zoho[query_date]:
        print(f"{amt}")
        total += amt
    print(f"Zoho Total({query_date}): {total}")
    

bank.close()