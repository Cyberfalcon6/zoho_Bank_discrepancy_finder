import openpyxl
from datetime import datetime
import collections
def generate_zoho_report():
    zoho = openpyxl.load_workbook(filename="inputs/zoho_transactions.xlsx")
    bs = openpyxl.load_workbook(filename="inputs/bank_statement.xlsx")

    zoho_sheet = zoho['sheet1']
    bs_sheet = bs['s']

    zoho_lost = openpyxl.load_workbook(filename="zoho_lost.xlsx")
    zoho_found_sheet = zoho_lost['found']
    zoho_found_different_date_sheet = zoho_lost['found_different_period']
    zoho_lost_sheet = zoho_lost['lost']
    zoho_transactions = collections.defaultdict(list)
    bs_transactions = collections.defaultdict(list)


    zoho_found_row = 1
    zoho_lost_row = 1
    zoho_found_different_date_row = 1


    n_transactions = 0

    # putting all bs transactions in a dictionary
    for row in bs_sheet.iter_rows(2, bs_sheet.max_row):
        amount = row[4].value
        bs_transactions[amount].append({'Tx date':row[1].value, 'Value Date': row[2].value, 'Description': row[3].value, 'Amount': amount})


    # checking if each zoho transaction amount exist in bank statement and if it's the same month
    for row in zoho_sheet.iter_rows(2, zoho_sheet.max_row):
        zoho_amount = row[4].value  #zoho amount
        zoho_date = row[0].value  
        if(not zoho_date):
            print(f"--------------- Reached The end of the sheet!! ¯\\_( ツ )_/¯  ---------------- {n_transactions} Transactions") 
            break
        current_month = zoho_date.month
        this_transaction = bs_transactions.get(zoho_amount)
        if(zoho_amount in bs_transactions):  # If this zoho amount was found in Bank statement 
            found = False

            temporary_bank_tx = None

            for bank_tx in bs_transactions[zoho_amount][:]:  # looking for all bank statement transactions under the this zoho amount
                temporary_bank_tx = bank_tx
                bank_tx_month = bank_tx['Tx date'].month
                print(f"Bank: {bank_tx_month} Zoho: {current_month} {int(bank_tx_month) == int(current_month)}")
                if(int(bank_tx_month) == int(current_month)): # If any of the bank statement transactions is in the current month
                    zoho_found_sheet[f"A{zoho_found_row}"] = row[0].value
                    zoho_found_sheet[f"B{zoho_found_row}"] = row[1].value
                    zoho_found_sheet[f"C{zoho_found_row}"] = row[2].value
                    zoho_found_sheet[f"D{zoho_found_row}"] = row[3].value
                    zoho_found_sheet[f"E{zoho_found_row}"] = row[4].value


                    zoho_found_sheet[f"H{zoho_found_row}"] = bank_tx['Tx date']
                    zoho_found_sheet[f"I{zoho_found_row}"] = bank_tx['Value Date']
                    zoho_found_sheet[f"J{zoho_found_row}"] = bank_tx['Description']
                    zoho_found_sheet[f"K{zoho_found_row}"] = bank_tx['Amount']
                    zoho_found_row += 1
                    bs_transactions[zoho_amount].remove(bank_tx)
                    found = True
                    break
            if(not found):
                if(temporary_bank_tx):
                    zoho_found_different_date_sheet[f"A{zoho_found_different_date_row}"] = row[0].value
                    zoho_found_different_date_sheet[f"B{zoho_found_different_date_row}"] = row[1].value
                    zoho_found_different_date_sheet[f"C{zoho_found_different_date_row}"] = row[2].value
                    zoho_found_different_date_sheet[f"D{zoho_found_different_date_row}"] = row[3].value
                    zoho_found_different_date_sheet[f"E{zoho_found_different_date_row}"] = row[4].value
                    print(temporary_bank_tx)
                    zoho_found_different_date_sheet[f"H{zoho_found_different_date_row}"] = temporary_bank_tx['Tx date']
                    zoho_found_different_date_sheet[f"I{zoho_found_different_date_row}"] = temporary_bank_tx['Value Date']
                    zoho_found_different_date_sheet[f"J{zoho_found_different_date_row}"] = temporary_bank_tx['Description']
                    zoho_found_different_date_sheet[f"K{zoho_found_different_date_row}"] = temporary_bank_tx['Amount']
                    zoho_found_different_date_row += 1
        else: 
            zoho_lost_sheet[f"A{zoho_lost_row}"] = row[0].value
            zoho_lost_sheet[f"B{zoho_lost_row}"] = row[1].value
            zoho_lost_sheet[f"C{zoho_lost_row}"] = row[2].value
            zoho_lost_sheet[f"D{zoho_lost_row}"] = row[3].value
            zoho_lost_sheet[f"E{zoho_lost_row}"] = row[4].value

            zoho_lost_row += 1
        n_transactions +=1



    zoho_lost.save("outputs/zoho_lost.xlsx")

if __name__ == '__main__':
    generate_zoho_report()