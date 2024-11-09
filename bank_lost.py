import openpyxl as op
import collections
from datetime import datetime
def generate_bank_report():
    bs = op.load_workbook("bank_statement.xlsx")
    zoho = op.load_workbook("zoho_transactions.xlsx")
    lost = op.load_workbook("lost.xlsx")


    zoho_sheet = zoho['sheet1']
    lost_sheet = lost['lost']
    found_sheet = lost['found']
    found_different_period_sheet = lost['found_different_period']
    bs_sheet = bs['s']

    column = 0
    row = 1

    transactions_found = 0

    lost_row = 1
    found_row = 1
    found_different_period_row = 1
    n_transactions = 0
    zoho_transactions = collections.defaultdict(list)

    for row in zoho_sheet.iter_rows(2, zoho_sheet.max_row):
        amount = row[4].value
        zoho_transactions[amount].append({'date':row[0].value, 'Reference': row[1].value, 'Type': row[2].value, 'Status': row[3].value, 'Amount': amount})


    for row in bs_sheet.iter_rows(2, bs_sheet.max_row):
        bs_amount = row[4].value  
        bs_date = row[1].value  


        if(isinstance(bs_date, str)): 
            current_month = bs_date.month
        elif(isinstance(bs_date, datetime)):
            current_month = bs_date.month
        elif(not bs_date):
            print(f"==============Reached the end of the sheet!!==================== \n Generated 'lost.xlsx' as a report {n_transactions}" )
            break
        else: 
            continue


        if(bs_amount in zoho_transactions):  # If this bs amount was found in zoho
            found = False
            temporary_zoho_tx = None
            for zoho_tx in zoho_transactions[bs_amount][:]:  # looking for all bank statement transactions under the this zoho amount
                temporary_zoho_tx = zoho_tx
                print(f"Bank: {current_month} Zoho: {int(zoho_tx['date'].month)} {int(zoho_tx['date'].month) == int(current_month)}")
                if(int(zoho_tx['date'].month) == int(current_month)): # If any of the bank statement transactions is in the current month
                    found_sheet[f"A{found_row}"] = row[0].value
                    found_sheet[f"B{found_row}"] = row[1].value
                    found_sheet[f"C{found_row}"] = row[2].value
                    found_sheet[f"D{found_row}"] = row[3].value
                    found_sheet[f"E{found_row}"] = row[4].value

                    found_sheet[f"H{found_row}"] = zoho_tx['date']
                    found_sheet[f"I{found_row}"] = zoho_tx['Reference']
                    found_sheet[f"J{found_row}"] = zoho_tx['Type']
                    found_sheet[f"K{found_row}"] = zoho_tx['Status']
                    found_sheet[f"L{found_row}"] = zoho_tx['Amount']
                    found_row += 1
                    zoho_transactions[bs_amount].remove(zoho_tx)
                    found = True
                    break
            if(not found):
                if(temporary_zoho_tx):
                    found_different_period_sheet[f"A{found_different_period_row}"] = row[0].value
                    found_different_period_sheet[f"B{found_different_period_row}"] = row[1].value
                    found_different_period_sheet[f"C{found_different_period_row}"] = row[2].value
                    found_different_period_sheet[f"D{found_different_period_row}"] = row[3].value
                    found_different_period_sheet[f"E{found_different_period_row}"] = row[4].value
                    found_different_period_sheet[f"H{found_different_period_row}"] = temporary_zoho_tx['date']
                    found_different_period_sheet[f"I{found_different_period_row}"] = temporary_zoho_tx['Reference']
                    found_different_period_sheet[f"J{found_different_period_row}"] = temporary_zoho_tx['Type']
                    found_different_period_sheet[f"K{found_different_period_row}"] = temporary_zoho_tx['Status']
                    found_different_period_sheet[f"L{found_different_period_row}"] = temporary_zoho_tx['Amount']
                    found_different_period_row += 1
        else: 
            lost_sheet[f"A{lost_row}"] = row[0].value
            lost_sheet[f"B{lost_row}"] = row[1].value
            lost_sheet[f"C{lost_row}"] = row[2].value
            lost_sheet[f"D{lost_row}"] = row[3].value
            lost_sheet[f"E{lost_row}"] = row[4].value

            lost_row += 1
        n_transactions += 1


    lost.save("lost.xlsx")
if __name__ == '__main__':
    generate_bank_report()
