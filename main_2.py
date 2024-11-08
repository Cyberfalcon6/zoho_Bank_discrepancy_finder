import openpyxl as op

# Load files
bs = op.load_workbook("bank_statement.xlsx")
zoho = op.load_workbook("zoho_transactions.xlsx")
lost = op.load_workbook("enhanced_lost.xlsx")
zoho_sheet = zoho['sheet1']
bs_sheet = bs['sheet1']
found_sheet = lost['found']
found_different_period_sheet = lost['found_different_period']
lost_sheet = lost['lost']

def copy_data(sheet, row, zoho_row, bs_row=None):
    for col in 'ABCDE':
        sheet[f"{col}{row}"] = zoho_sheet[f"{col}{zoho_row}"].value
    if bs_row:
        for col, offset in zip('IJKL', range(1, 5)):
            sheet[f"{col}{row}"] = bs_sheet[f"{chr(64+offset)}{bs_row}"].value
    return row + 1

# Convert sheets into dictionaries
bs_data = {}
for b in range(2, bs_sheet.max_row + 1):
    amount = int(bs_sheet[f"E{b}"].value)
    date = str(bs_sheet[f"B{b}"].value)
    month = date.split("/")[1]
    bs_data.setdefault(amount, []).append((date, month, b))

zoho_data = {}
for r in range(2, zoho_sheet.max_row + 1):
    amount = int(zoho_sheet[f"E{r}"].value)
    date = str(zoho_sheet[f"A{r}"].value)
    month = date.split("/")[1]
    zoho_data.setdefault(amount, []).append((date, month, r))

lost_row = found_row = found_different_period_row = 1

# Find matches
for amount, bs_transactions in bs_data.items():
    if amount in zoho_data:
        for bs_date, bs_month, b_row in bs_transactions:
            match_found = False
            for zoho_date, zoho_month, z_row in zoho_data[amount]:
                if bs_date == zoho_date:
                    # Record in found sheet
                    match_found = True
                    found_row = copy_data(found_sheet, found_row, z_row, b_row)
                elif bs_month == zoho_month:
                    # Record in found_different_period sheet
                    match_found = True
                    found_different_period_row = copy_data(found_different_period_sheet, found_different_period_row, z_row, b_row)
            if not match_found:
                # Record in lost sheet
                lost_row = copy_data(lost_sheet, lost_row, b_row, None)

# Save
lost.save("enhanced_lost.xlsx")
print("Thank you for using our program!")


